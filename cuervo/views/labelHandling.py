from django.shortcuts import render, redirect, get_object_or_404, reverse
from ..models import label, coil, labelStatus, init_label, order, lot, coilStatus, coil_request, \
    coil_request_status, coilType, inventoryLocation, granel_lot, SKU, zip_file_parent, zip_file_child, log_files, sku_Type, coilProvider, sku_SubType 
from django.views.generic import UpdateView
from django.urls import reverse_lazy
from ..form import FilterLabelForm, UpdateLabelForm, LabelInitForm, CreateCoilFormv2, LabelInitFormCSV, LabelInitInventoryForm
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
import csv
from datetime import datetime, timedelta
from django.contrib import messages
from openpyxl import load_workbook
import re
from django.utils import timezone
from django.http import HttpResponse
import pandas as pd
import pytz
from jose_cuervo.settings import TIME_ZONE
from django.core.paginator import Paginator
from django.http import JsonResponse
from urllib.parse import unquote
from django.http import JsonResponse
import zipfile
import os
import pyzipper
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from django.db.models import Q
from django.db import transaction

def setSystemTimeZoneToDatetime(someDateTime):
    # Expects to recieve naive datetime with a local datetime value. Then this function includes system time to
    # convert to UTC
    if (someDateTime.tzinfo is None):
        tz = pytz.timezone(TIME_ZONE)
        if (type(someDateTime) is pd.Timestamp):
            someDt = pd.date_range(start=someDateTime, periods=1).tz_localize(tz)
            newDate = someDt[0]
        else:
            newDate = someDateTime.astimezone(tz)
    else:
        newDate = someDateTime

    return newDate


def autocomplete_sku(request):
    if 'term' in request.GET:
        qs = SKU.objects.filter(sku__icontains=request.GET.get('term'))
        titles = list(qs.values('sku'))
        return JsonResponse(titles, safe=False)
    return JsonResponse([], safe=False)
    
def autocomplete_subbrand(request):
    if 'term' in request.GET:
        qs = sku_SubType.objects.filter(name__icontains=request.GET.get('term'))
        titles = list(qs.values('name'))
        return JsonResponse(titles, safe=False)
    return JsonResponse([], safe=False)
    
def autocomplete_brand(request):
    if 'term' in request.GET:
        # Filtrar por name o description que contengan el término de búsqueda
        qs = sku_Type.objects.filter(
            Q(name__icontains=request.GET.get('term')) |
            Q(description__icontains=request.GET.get('term'))
        )
        # Concatenar name y description para mostrar en las sugerencias
        suggestions = list(qs.values_list('name', 'description'))
        results = [{'label': f"{name}, {description}", 'value': f"{name}, {description}"} for name, description in suggestions]
        return JsonResponse(results, safe=False)
    return JsonResponse([], safe=False)

@permission_required('cuervo.view_label', login_url='/login/')
def labelHandling_view(request):
    label_list = None
    page_obj = None
    now = datetime.now()
    form = FilterLabelForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            request.session['startDatetime'] = request.POST.get('startDatetime')
            request.session['endDatetime'] = request.POST.get('endDatetime')
            request.session['sku'] = form.cleaned_data.get('sku') if form.cleaned_data.get('sku') else None
    else:
        form.fields['startDatetime'].initial = (now - timedelta(days=1)).strftime('%Y-%m-%d %H:%M')
        form.fields['endDatetime'].initial = now.strftime('%Y-%m-%d %H:%M')

    startdatetime = request.session.get('startDatetime')
    enddatetime = request.session.get('endDatetime')
    sku = request.session.get('sku')

    if startdatetime and enddatetime:
        startdatetimeUTC = setSystemTimeZoneToDatetime(datetime.strptime(startdatetime, '%Y-%m-%d %H:%M'))
        enddatetimeUTC = setSystemTimeZoneToDatetime(datetime.strptime(enddatetime, '%Y-%m-%d %H:%M'))

        labels = label.objects.filter(last_update__range=[startdatetimeUTC, enddatetimeUTC])

        if sku:
            coil_list = coil.objects.filter(sku=sku)
            labels = labels.filter(FK_coil_id__in=coil_list)

        paginator = Paginator(labels, 100)  # Muestra 100 registros por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        label_list = page_obj

    return render(request, 'cuervo/labelHandling.html', {'form': form, 'labelList': label_list, 'page_obj': page_obj})

class updateLabel_view(PermissionRequiredMixin, UpdateView):
    model = label
    template_name = 'cuervo/label_edit.html'
    success_url = reverse_lazy('labelHandling')
    form_class = UpdateLabelForm
    permission_required = 'cuervo.change_label'

#<----------------------- Search Label By Coil ID -------------------------------->
def searchLabelByCoilFK(request, pk):
    coilFk = get_object_or_404(coil, id=pk)
    labels = label.objects.filter(FK_coil_id=coilFk)
    coil_not_delivered = coilFk.notDelivered

    if request.method == 'POST':
        selected_labels = request.POST.getlist('selected_labels')
        faltante = request.POST.get('faltante', 'off')

        selected_labels_count = len(selected_labels)
        faltante_labels_count = labels.filter(FK_labelStatus_id__name='Faltante').count()

        if faltante == 'on':
            faltante_labels_count += selected_labels_count

        if faltante_labels_count <= coil_not_delivered:
            for label_id in selected_labels:
                label_obj = label.objects.get(id=label_id)
                if faltante == 'on':
                    label_obj.FK_labelStatus_id = labelStatus.objects.get(name='Faltante')
                else:
                    label_obj.FK_labelStatus_id = labelStatus.objects.get(name='Faltante')  # Define your other status
                label_obj.save()

        # Redirect to the same page to avoid form resubmission
        return redirect('label-find', pk=pk)

    total_labels = labels.count()
    faltante_labels = labels.filter(FK_labelStatus_id__name='Faltante').count()

    disable_checkboxes = faltante_labels >= coil_not_delivered

    return render(request, 'cuervo/label_by_CoilFk.html', {
        'labels': labels,
        'disable_checkboxes': disable_checkboxes,
    })

#-------------------- Init Label  ------------------------------

def progress_charge_labels(request):
    return render(request, "cuervo/progress_charge_labels.html")

def get_unseen_files_count(request):
    unseen_files_count = zip_file_parent.objects.filter(seen=False).count()
    return {'unseen_files_count': unseen_files_count}

@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def mark_as_seen(request, id):
    if request.method == 'POST':
        try:
            parent_zip = zip_file_parent.objects.get(id=id)
            parent_zip.seen = True
            parent_zip.save()
            # Prepara el mensaje de éxito
            message = f'{parent_zip.file_name_parent} marcado como visto.'
            return JsonResponse({'success': True, 'message': message})
        except zip_file_parent.DoesNotExist:
            message = 'El archivo especificado no existe.'
            return JsonResponse({'success': False, 'message': message})
    else:
        message = 'Solicitud inválida.'
        return JsonResponse({'success': False, 'message': message})
        
def view_log_files(request, id):
    child_zips = zip_file_child.objects.filter(parent_id=id)
    log_files_list = log_files.objects.filter(FK_zip_child__in=child_zips)
    context = {
        'log_files_list': log_files_list,
    }
    return render(request, 'cuervo/log_files_table.html', context)

def view_folios(request, file_name):
    csv_file_path = f'E:\\ministration_files\\csv-files\\{file_name}'
    init_labels = init_label.objects.filter(file_name=csv_file_path)

    # Paginación
    paginator = Paginator(init_labels, 50)  # Muestra 50 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'init_labels': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'cuervo/registed_labels_csv.html', context)

def progress_data_ajax(request):
    # Obtener todos los objetos zip_file_parent
    parents = zip_file_parent.objects.filter(seen=False)

    # Crear una lista de progreso
    progress_data = []
    for parent in parents:
        total_files = parent.num_files
        processed_files = parent.num_processed_files
        progress_percentage = (processed_files / total_files) * 100 if total_files > 0 else 0

        # Verificar si existen log_files asociados a este parent
        child_zips = zip_file_child.objects.filter(parent=parent)
        has_log_files = log_files.objects.filter(FK_zip_child__in=child_zips).exists()

        progress_data.append({
            'id': parent.id,
            'file_name_parent': parent.file_name_parent,
            'progress_percentage': progress_percentage,
            'processed_files': processed_files,
            'total_files': total_files,
            'has_log_files': has_log_files,  # Nuevo campo
        })

    return JsonResponse({'progress_data': progress_data})
    
def charge_csv_server(request):
    msg = None
    if request.method == "POST":
        form = LabelInitFormCSV(request.POST, request.FILES)
        if form.is_valid():
            csv_files = request.FILES.getlist('csv_file')
            brand = form.cleaned_data.get('brand')
            ministrationNumber = form.cleaned_data.get('ministrationNumber')

            # Usar la ruta especificada
            csv_destination_dir = r'E:\ministration_files\csv-files'

            # Crear el directorio si no existe
            os.makedirs(csv_destination_dir, exist_ok=True)
            processed_files = []
            child_objects = []
            try:
                brand_name, brand_description = map(str.strip, brand.split(','))    
                brand_obj = sku_Type.objects.filter(name=brand_name, description=brand_description).first()
                for file in csv_files:
                    if not file.name.endswith('.csv'):
                        messages.error(request, 'Este archivo no contiene la extension .csv')
                        break
                    # Guardar el archivo CSV en la carpeta especificada
                    file_path = os.path.join(csv_destination_dir, file.name)
                    with open(file_path, 'wb+') as destination:
                        for chunk in file.chunks():
                            destination.write(chunk)

                    # Crear una instancia del modelo para el archivo subido

                    child_obj = zip_file_child.objects.create(
                        brand_name=brand_obj,
                        ministration_number=ministrationNumber,
                        file_name=file.name,
                    )
                    child_objects.append(child_obj)
                    processed_files.append(file.name)

                if not msg:
                    messages.success(request, 'Los archivos CSV se han subido correctamente.')

                # Guardar la referencia en el modelo parent (si es necesario)
                if processed_files:
                    for child_obj in child_objects:
                        zip_file_obj = zip_file_parent.objects.create(
                            brand_name=brand_obj,
                            ministration_number=ministrationNumber,
                            file_name_parent=child_obj.file_name,
                            num_files=1
                        )
                        child_obj.parent = zip_file_obj
                        child_obj.save()
            except Exception as e:
                messages.error(request, 'Verifica que toda la información sea correcta')
                print(str(e))
        else:
            msg = 'Ha ocurrido un error en el formulario'
            print(form.errors)
    else:
        form = LabelInitForm()

    return render(request, "cuervo/label_init_create_csv.html", {"form": form, "msg": msg})

def charge_zip_server(request):
    msg = None
    corrupted_files_list = []
    coil_msg_list = []
    if request.method == "POST":
        form = LabelInitForm(request.POST, request.FILES)
        if form.is_valid():
            csv_files = request.FILES.getlist('csv_file')
            brand = form.cleaned_data.get('brand')
            zip_pass = form.cleaned_data.get('zip_pass')
            ministrationNumber = form.cleaned_data.get('ministrationNumber')
            
             # Usar las rutas especificadas
            destination_dir = r'E:\ministration_files\zip-files'
            csv_destination_dir = r'E:\ministration_files\csv-files'

            # Crear los directorios si no existen
            os.makedirs(destination_dir, exist_ok=True)
            os.makedirs(csv_destination_dir, exist_ok=True)
            processed_files = []
            child_objects = []
            counter = 0

            try:
                for file in csv_files:
                    if not file.name.endswith('.zip'):
                        continue
                    brand_name, brand_description = map(str.strip, brand.split(','))    
                    brand_obj = sku_Type.objects.filter(name=brand_name, description=brand_description).first()
                    with zipfile.ZipFile(file, 'r') as zip_ref:
                        test_file_names = zip_ref.namelist()  # Obtener los nombres de los archivos en el zip
                        processed_files.extend(
                            test_file_names)  # Guardar los nombres de archivos para su posterior referencia
                        for zip in test_file_names:
                            zip_ref.extract(zip, path=destination_dir)

                        # Procesar solo los archivos ZIP que se cargaron previamente
                        for root, dirs, files in os.walk(destination_dir):
                            for filename in files:
                                file_path = os.path.join(root, filename)
                                if filename in processed_files:
                                    if not file_path.endswith('.zip'):
                                        os.remove(file_path)
                                        continue

                                    try:
                                        with pyzipper.AESZipFile(file_path, 'r',
                                                                 compression=pyzipper.ZIP_DEFLATED) as extracted_zip:
                                            for extracted_file in extracted_zip.namelist():
                                                try:
                                                    extracted_zip.extract(extracted_file, path=csv_destination_dir,
                                                                          pwd=zip_pass.encode('utf-8'))
                                                    child_obj = zip_file_child.objects.create(
                                                        brand_name=brand_obj,
                                                        ministration_number=ministrationNumber,
                                                        file_name=extracted_file,
                                                    )
                                                    child_objects.append(child_obj)
                                                    counter += 1
                                                except RuntimeError as e:
                                                    if 'Bad password for file' in str(e):
                                                        msg = f'Contraseña incorrecta para los archivos'
                                                    else:
                                                        msg = f'Error inesperado al procesar los archivos: {str(e)}'
                                                    # Eliminar el archivo ZIP original en caso de error
                                                    if os.path.exists(file_path):
                                                        os.remove(file_path)
                                    except Exception as e:
                                        msg = 'a'
                                        # Eliminar el archivo ZIP original en caso de error al abrirlo
                                        corrupted_files_list.append(filename)
                                        if os.path.exists(file_path):
                                            os.remove(file_path)
                                            
                    # Eliminar el archivo ZIP después de procesarlo
                    if os.path.exists(os.path.join(destination_dir, file.name)):
                        os.remove(os.path.join(destination_dir, file.name))

                    zip_file_obj = zip_file_parent.objects.create(
                        brand_name=brand_obj,
                        ministration_number=ministrationNumber,
                        file_name_parent=file.name,
                        password=zip_pass,
                        num_files=counter
                    )

                    for child_obj in child_objects:
                        child_obj.parent = zip_file_obj
                        child_obj.save()

                if not msg:
                   messages.success(request, 'La contraseña es correcta y los archivos zip se pueden abrir')
                elif msg != 'a':
                   messages.error(request, msg)   
                
                print(corrupted_files_list)
                if corrupted_files_list:
                    for corrupted_file in corrupted_files_list:
                        coil_msg_list.append(f"El archivo {corrupted_file} no se pudo abrir porque está corrupto.<br/>")
                    error_message = ' '.join(coil_msg_list)
                    print(f'Se ha hecho la carga pero hubo un error al procesar los siguientes archivos: <br/>{error_message}')
                    messages.warning(request, f'Se ha hecho la carga pero hubo un error al procesar los siguientes archivos: <br/>{error_message}')



            except zipfile.BadZipFile:
                msg = 'El archivo no es un archivo zip válido'
            except RuntimeError as e:
                if 'Bad password for file' in str(e):
                    msg = 'Contraseña incorrecta para el archivo zip'
                else:
                    msg = f'Error inesperado: {str(e)}'
            except Exception as e:
                msg = f'Error inesperado: {str(e)}'
                print(msg)
        else:
            msg = 'Ha ocurrido un error en el formulario'
            print(form.errors)
    else:
        form = LabelInitForm()

    return render(request, "cuervo/label_init_create.html", {"form": form, "msg": msg})

@permission_required('cuervo.add_labelstatus', login_url='/login/')
def init_label_information(request):
        msg = None
        issaved = None
        if request.method == "POST":
            form = LabelInitForm(request.POST)
            if form.is_valid():
                csv_files = request.FILES.getlist('csv_file')
                brand = form.cleaned_data.get('brand')
                # Obtener la fecha actual
                fecha_actual = datetime.now()

                # Sumar nueve meses a la fecha actual
                anno = fecha_actual.year + (fecha_actual.month + 9) // 12
                mes = (fecha_actual.month + 9) % 12
                if mes == 0:
                    mes = 12
                dia = min(fecha_actual.day, [31,
                                             29 if anno % 4 == 0 and (anno % 100 != 0 or anno % 400 == 0) else 28,
                                             31, 30, 31, 30, 31, 31, 30, 31, 30, 31][mes - 1])

                fecha_nueve_meses_despues = datetime(anno, mes, dia)
                ministrationNumber = form.cleaned_data.get('ministrationNumber')
                for csv_file in csv_files:
                    if not csv_file.name.endswith('.csv'):
                        # Handle error: Invalid file format for a specific file
                        continue  # Skip this file and move to the next one

                    # Columnas a excluir (ID y Registro en este caso)
                    columns_to_exclude = ['ID', 'REGISTRO']

                    folios = []
                    textos = []

                    decoded_file = csv_file.read().decode('utf-8')
                    csv_reader = csv.reader(decoded_file.splitlines(), delimiter=',')

                    # Obtiene los encabezados del archivo CSV
                    headers = next(csv_reader)

                    # Encuentra los índices de las columnas a excluir
                    exclude_indices = [headers.index(column) for column in columns_to_exclude if column in headers]

                    for row in csv_reader:
                        # Excluye las columnas en cada fila
                        filtered_row = [value for index, value in enumerate(row) if index not in exclude_indices]
                        # Divide los folios y los textos en arreglos separados
                        folios.extend(filtered_row[::2])  # Obtén los folios (columnas pares)
                        textos.extend(filtered_row[1::2])  # Obtén los textos (columnas impares)

                    # Verifica si hay contenido repetido en los textos
                    duplicates = []
                    seen = set()
                    for i, texto in enumerate(textos):
                        if texto in seen:
                            duplicates.append(i)
                        else:
                            seen.add(texto)

                    folios_filtrado = [valor for valor in folios if not valor.isdigit()]
                    if (textos != []) and folios_filtrado != [] and not duplicates:
                        data_exists = label.objects.filter(uniqueid__in=textos, url__in=folios_filtrado)
                        folios_filtrado = [valor for valor in folios_filtrado if valor.strip()]
                        if not data_exists:
                            try:
                                if len(folios_filtrado) == len(textos):
                                    for index, x in enumerate(folios_filtrado):
                                         init_label.objects.create(
                                            uniqueid=folios_filtrado[index],
                                            url=textos[index],
                                            file_name=str(csv_file.name),
                                            brand=brand,
                                            ministrationNumber=ministrationNumber,
                                            expiration=fecha_nueve_meses_despues
                                         ).save()
                            except Exception as e:
                                #label.objects.filter(FK_coil_id=coilObj).delete()
                                messages.error(request, "Error al generar marbetes" + str(e))
                        else:
                            messages.error(request, "Algunos de estos marbetes ya existen")
                    else:
                        if duplicates:
                            # Aquí puedes trabajar con los datos repetidos
                            for index in duplicates:
                                print(f"El texto '{textos[index]}' está repetido en el folio: {folios[index]}.")
                        else:
                            messages.error(request, 'Revisa si los números de folio o los folios no entregados estén bien')
                issaved = True
            else:
                msg = 'Ha ocurrido un error'
                print(form.errors)
        else:
            form = LabelInitForm()
        if issaved:
            messages.success(request, 'La información ha sido procesada correctamente')

        return render(request, "cuervo/label_init_create.html", {"form": form, "msg": msg})


def extract_data_from_excel_No_Rollo(file, sheet_name_pattern):
    values = []
    # Load the Excel workbook
    wb = load_workbook(file)

    # Find the worksheet that matches the pattern
    matching_sheets = [ws for ws in wb.sheetnames if re.search(sheet_name_pattern, ws, re.IGNORECASE)]
    if not matching_sheets:
        raise ValueError("No sheet found matching the pattern: {}".format(sheet_name_pattern))

    ws = wb[matching_sheets[0]]
    row = ws.max_row
    for i in range(2, row + 1):
        cell_value = ws.cell(row=i, column=1).value
        if cell_value is not None:
            values.append(cell_value)
        else:
            break  # Stop loop if encountering a None value

    return values

def extract_data_from_excel_folios(file, sheet_name_pattern):
    values = []
    # Load the Excel workbook
    wb = load_workbook(file)

    # Find the worksheet that matches the pattern
    matching_sheets = [ws for ws in wb.sheetnames if re.search(sheet_name_pattern, ws, re.IGNORECASE)]
    if not matching_sheets:
        raise ValueError("No sheet found matching the pattern: {}".format(sheet_name_pattern))

    ws = wb[matching_sheets[0]]
    row = ws.max_row
    #print(ws['A1'].value)
    for i in range(2, row + 1):
        cell_value = ws.cell(row=i, column=2).value
        if cell_value is not None:
            values.append(cell_value)
        else:
            break  # Stop loop if encountering a None value

    return values

@permission_required('cuervo.add_labelstatus', login_url='/login/')
def init_label_canceled(request):
        msg = None
        issaved = None
        if request.method == "POST":
            csv_files = request.FILES.getlist('csv_file')
            for file in csv_files:
                sheet_name = 'folios con defecto'
                # Extract data from the specified cell
                valueNoRollo = extract_data_from_excel_No_Rollo(file, sheet_name)
                valueFolios = extract_data_from_excel_folios(file, sheet_name)

        if issaved:
            return redirect('/labelMenu/')

        return render(request, "cuervo/label_canceled.html", {"msg": msg})



def extract_data_from_excel_defect_label(file, sheet_name_pattern, request):
    values = {}
    no_rollo = []
    folios = []
    num_caja = []

    # Load the Excel workbook with data_only=True
    wb = load_workbook(file, data_only=True)

    # Find the worksheet that matches the pattern
    matching_sheets = [ws for ws in wb.sheetnames if re.search(sheet_name_pattern, ws, re.IGNORECASE)]
    if not matching_sheets:
        messages.warning(request, 'No se encontró la hoja llamada "Folios con defecto" dentro del excel')
    else:
        ws = wb[matching_sheets[0]]

        # Iterate over rows starting from row 13
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            if any(cell is None for cell in row):
                break
            no_rollo.append(row[0])
            folios.append(row[1])
            num_caja.append(row[2])

        values['folios'] = folios
        values['no_rollo'] = no_rollo
        values['num_caja'] = num_caja
    return values

def extract_data_from_excel_warehouse_label(file, sheet_name_pattern, request):
    values = {}
    no_rollo = []
    folios_iniciales = []
    folios_finales = []
    folios_utilizados = []
    folios_faltantes = []
    cantidad_rollo = []
    num_caja = []
    cantidad_caja = []

    # Load the Excel workbook with data_only=True
    wb = load_workbook(file, data_only=True)

    # Find the worksheet that matches the pattern
    matching_sheets = [ws for ws in wb.sheetnames if re.search(sheet_name_pattern, ws, re.IGNORECASE)]
    if not matching_sheets:
        messages.error(request, 'No se encontró la hoja llamada "OT" dentro del excel')
    else:
        ws = wb[matching_sheets[0]]

        values['proveedor'] = ws['B4'].value
        values['ot'] = ws['B7'].value
        values['etiqueta'] = ws['B8'].value
        values['item'] = ws['B9'].value

        values['orden'] = ws['D7'].value
        values['desc'] = ws['D8'].value
        values['factura'] = ws['D9'].value

        # Create a dictionary to store the values for merged cells
        merged_cell_ranges = ws.merged_cells.ranges

        # Iterate over rows starting from row 13
        for row in ws.iter_rows(min_row=13, min_col=1, max_col=8, values_only=False):
            row_values = []

            for cell in row:
                # Check if the cell is part of a merged range
                is_merged = any(cell.coordinate in merge for merge in merged_cell_ranges)
                if is_merged:
                    # Get the top-left cell of the merged range to use its value
                    top_left_cell = next((merge.start_cell for merge in merged_cell_ranges if cell.coordinate in merge), cell)
                    cell_value = top_left_cell.value
                else:
                    # Use the cell's own value
                    cell_value = cell.value

                row_values.append(cell_value)

            # Check if the row contains only None values (stop condition)
            if all(value is None for value in row_values):
                break

            # Append the values to the respective lists
            try:
                numrollo = int(row_values[0])
                cr = int(row_values[5])
                no_rollo.append(row_values[0])
                folios_iniciales.append(row_values[1])
                folios_finales.append(row_values[2])
                folios_utilizados.append(row_values[3])
                folios_faltantes.append(row_values[4])
                cantidad_rollo.append(row_values[5])
                num_caja.append(row_values[6])
                cantidad_caja.append(row_values[7])
            except Exception as e:
                messages.error(request, f"No se ha cargado la bobina con el rango {row_values[1]} - {row_values[2]} la causa puede estar en que el No. de Caja o No. de Rollo no sea un número entero o el campo este vacio")

        if "Total" in str(folios_finales[-1]):
            no_rollo.pop()
            folios_iniciales.pop()
            folios_finales.pop()
            folios_utilizados.pop()
            folios_faltantes.pop()
            cantidad_rollo.pop()
            num_caja.pop()
            cantidad_caja.pop()

        # Assign the lists to the values dictionary
        values['folios_iniciales'] = folios_iniciales
        values['folios_finales'] = folios_finales
        values['folios_utilizados'] = folios_utilizados
        values['folios_faltantes'] = folios_faltantes
        values['cantidad_rollo'] = cantidad_rollo
        values['num_caja'] = num_caja
        values['cantidad_caja'] = cantidad_caja
        values['num_rollo'] = no_rollo

    return values

@permission_required('auth.load_coil', login_url='/login/')
def init_label_in_inventory(request):
    msg = None
    coil_msg_list = []
    label_msg_list = []
    coil_filter = None
    label_filter = None
    loaded_labels = None
    form = LabelInitInventoryForm()
    
    if request.method == "POST":
        form = LabelInitInventoryForm(request.POST, request.FILES)
        if form.is_valid():
            brand = form.cleaned_data.get('brand')
            csv_files = request.FILES.getlist('csv_file')

            obj_status = coilStatus.objects.get(name='Disponible')
            obj_type = coilType.objects.all().first()
            obj_label_status = labelStatus.objects.get(name='Asignado')
            obj_inventory = inventoryLocation.objects.get(name='Almacen')
            obj_damaged = labelStatus.objects.get(name='Dañados Proveedor')
            
            try:
                sku_brand = SKU.objects.filter(sku=brand).first()
                for file in csv_files:
                    sheet_name = 'OT'
                    # Extract data from the specified cell
                    labels = extract_data_from_excel_warehouse_label(file, sheet_name, request)
                    if labels['folios_iniciales'] and labels['folios_finales'] and labels['num_rollo'] and\
                            labels['folios_faltantes'] and labels['cantidad_rollo'] and labels['folios_utilizados']:
                        if (len(labels['folios_iniciales']) == len(labels['folios_finales']) ==
                                len(labels['num_rollo']) == len(labels['folios_faltantes']) ==
                                len(labels['cantidad_rollo']) == len(labels['folios_utilizados'])):
                            # Lista para hacer bulk_create de labels
                            labels_to_create = []
                            for i in range(len(labels['folios_iniciales'])):
                                if labels['folios_iniciales'][i] is not None and labels['folios_finales'][i] is not None \
                                        and labels['num_rollo'][i] is not None and labels['folios_faltantes'][i] is not None \
                                        and labels['cantidad_rollo'][i] is not None and labels['folios_utilizados'][i] is not None \
                                        and labels['num_caja'][i] is not None and labels['proveedor'] is not None \
                                        and labels['item'] is not None and labels['orden'] is not None:

                                    if labels['folios_iniciales'][i] < labels['folios_finales'][i]:
                                        if sku_brand:
                                            coil_filter = coil.objects.filter(
                                                initNumber=labels['folios_iniciales'][i],
                                                finishNumber=labels['folios_finales'][i],
                                                numrollo=labels['num_rollo'][i],
                                            )
                                            if not coil_filter:
                                                obj_coil = coil.objects.create(
                                                    initNumber=labels['folios_iniciales'][i],
                                                    finishNumber=labels['folios_finales'][i],
                                                    numrollo=labels['num_rollo'][i],
                                                    notDelivered=labels['folios_faltantes'][i],
                                                    missing=labels['cantidad_rollo'][i],
                                                    delivered=labels['folios_utilizados'][i],
                                                    boxNumber=labels['num_caja'][i],
                                                    qty_box=labels['cantidad_caja'][i],
                                                    purchaseOrder=labels['orden'],
                                                    orderUniqueid=labels['ot'],
                                                    sku=labels['item'],
                                                    FK_coilStatus_id=obj_status,
                                                    FK_coilType_id=obj_type,
                                                    last_edit_user=request.user,
                                                    FK_coilProvider_id=labels['proveedor'],
                                                    Fk_sku_subtype_id=sku_brand.Fk_sku_subtype_id
                                                )
                                                matching_labels = init_label.objects.filter(
                                                    uniqueid__gte=labels['folios_iniciales'][i],
                                                    uniqueid__lte=labels['folios_finales'][i]
                                                )
                                                print(matching_labels)
                                                if matching_labels:
                                                    for matching_label in matching_labels:
                                                        label_filter = label.objects.filter(url=matching_label.url)
                                                        obj_provider = coilProvider.objects.filter(number=labels['proveedor']).first()
                                                        if not label_filter:
                                                            # Recolectar los labels para hacer bulk_create
                                                            labels_to_create.append(label(
                                                                uniqueid=matching_label.uniqueid,
                                                                url=matching_label.url,
                                                                brand=matching_label.brand,
                                                                ministrationNumber=matching_label.ministrationNumber,
                                                                supplier=obj_provider,
                                                                FK_coil_id=obj_coil,
                                                                FK_labelStatus_id=obj_label_status,
                                                                FK_inventoryLocation_id=obj_inventory,
                                                                last_edit_user=request.user,
                                                                expiration=matching_label.expiration
                                                            ))
                                                        else:
                                                            mensaje_label = f"El marbete con la url {matching_label.url} ya ha sido registrada<br/>"
                                                            label_msg_list.append(mensaje_label)
                                                else:
                                                    messages.error(request, f"No se han cargado los marbetes asignados a esta bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]}")
                                                    loaded_labels = True
                                                    obj_coil.delete()
                                                              
                                            else:
                                                mensaje_coil = f"La bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} ya esta creada<br/>"
                                                coil_msg_list.append(mensaje_coil)
                                        else:
                                            messages.error(request,
                                                           f"No se ha cargado la bobina la causa puede estar en que el No. de Proveedor no sea valido")
                                    else:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina la causa puede estar en que los valores de los campos de Folio Inicial y Final esten invertidos")
                                else:
                                    if labels['item'] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en Item")
                              
                                    if labels['orden'] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en Orden de Compra")
                                    if labels['proveedor'] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en No. Proveedor")
                                    if labels['num_caja'][i] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en No. Caja")
                                    if labels['folios_faltantes'][i] is None:
                                        messages.error(request,
                                                       f"No se han cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en Folios Faltantes")
                                    if labels['folios_utilizados'][i] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en Folios Utilizados")
                                    if labels['cantidad_rollo'][i] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en Cantidad Rollo")
                                    if labels['num_rollo'][i] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina con el rango {labels['folios_iniciales'][i]} - {labels['folios_finales'][i]} la causa puede estar en un campo vacio en ID Rollo")
                                    if labels['folios_finales'][i] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina la causa puede estar en un campo vacio en Folios Finales")
                                    if labels['folios_iniciales'][i] is None:
                                        messages.error(request,
                                                       f"No se ha cargado la bobina la causa puede estar en un campo vacio en Folios Inicailes")
                            if labels_to_create:
                                label.objects.bulk_create(labels_to_create)
                            # Se lee el data sheet folios con defecto

                            # Se lee el data sheet folios con defecto
                            sheet = 'defecto'
                            labels_cancel = extract_data_from_excel_defect_label(file, sheet, request)

                            if labels_cancel['folios'] and labels_cancel['no_rollo'] and labels_cancel['num_caja']:

                                if len(labels_cancel['folios']) == len(labels_cancel['num_caja']) == len(labels_cancel['no_rollo']):

                                    for i in range(len(labels_cancel['folios'])):
                                        label_damaged = label.objects.get(
                                            uniqueid=labels_cancel['folios'][i],
                                            FK_labelStatus_id=obj_label_status,
                                            last_edit_user=request.user,
                                        )
                                        label_damaged.FK_labelStatus_id = obj_damaged
                                        label_damaged.save()

                        else:
                            messages.error(request, 'El numero de columnas no coinciden.')
                    else:
                        messages.error(request, 'No se encontraron todos los datos correspondientes en el archivo.')


                if coil_msg_list:
                    error_message = ' '.join(coil_msg_list)
                    messages.error(request, f'Error al cargar las bobinas: <br/>{error_message}')

                if label_msg_list:
                    error_message = ' '.join(label_msg_list)
                    messages.error(request, f'Error al cargar los marbetes:<br/> {error_message}')
                if not label_msg_list and not coil_msg_list and not loaded_labels:
                    messages.success(request, "Datos cargados exitosamente.")
                else:
                    messages.warning(request, "Revisar los errores en las bobinas indicadas, la demas información se cargo exitosamente")
            except FileNotFoundError as fnf_error:
                messages.error(request, f'Error al cargar el archivo: {str(fnf_error)}')
            except ValueError as value_error:
                messages.error(request, f'Error al procesar los datos del archivo: {str(value_error)}')
            except Exception as e:
                messages.error(request, f'Hubo un error al cargar datos: {str(e)}')
        else:
            msg = 'Ha ocurrido un error en el formulario'
            print(form.errors)
    return render(request, "cuervo/label_in_inventory.html", {"msg": msg, "form": form})


def view_coils(request, pk):
    solicitud = coil_request.objects.get(pk=pk)
    bobinas_ids = [int(id) for id in solicitud.requested_coils.split(',')]  # Obtener los IDs de las bobinas
    bobinas = coil.objects.filter(id__in=bobinas_ids)  # Obtener las instancias de las bobinas
    return render(request, 'cuervo/view_coils.html', {'solicitud': solicitud, 'bobinas': bobinas})

@permission_required('cuervo.view_coil_request', login_url='/login/')
def solicitudmarbete_request(request, pk):
    msg = None
    orden = None
    bobinas = []
    selected_lote_coils = []
    marbetes_necesarios = 0
    marbetes_totales = 0

    try:

        coil_request_instance = get_object_or_404(coil_request, pk=pk)

        if request.method == "POST":
            form = CreateCoilFormv2(request.POST)

            if 'crear' in request.POST and form.is_valid():
                bobinas_seleccionadas = request.POST.getlist('selected_bobinas')

                if not bobinas_seleccionadas:
                    msg = 'Debe seleccionar al menos una bobina para actualizar la solicitud.'
                else:
                    total_quantity = sum(
                        int(coil.missing) for coil in coil.objects.filter(id__in=bobinas_seleccionadas)
                    )

                    if not orden:
                        orden = coil_request_instance.FK_order_id

                    previously_selected_coils = [int(id) for id in coil_request_instance.requested_coils.split(
                        ',')] if coil_request_instance.requested_coils else []

                    coil_request_instance.FK_order_id = orden
                    coil_request_instance.requested_coils = ','.join(bobinas_seleccionadas)
                    coil_request_instance.request_date = timezone.now()
                    coil_request_instance.FK_coil_request_status_id = coil_request_status.objects.get(status='Pendiente')
                    coil_request_instance.created_by = request.user
                    coil_request_instance.total_number = total_quantity
                    coil_request_instance.save()

                    coil.objects.filter(id__in=bobinas_seleccionadas).update(
                        FK_coilStatus_id=coilStatus.objects.get(id=4)
                    )

                    bobinas_deseleccionadas = set(previously_selected_coils) - set(map(int, bobinas_seleccionadas))
                    if bobinas_deseleccionadas:
                        coil.objects.filter(id__in=bobinas_deseleccionadas).update(
                            FK_coilStatus_id=coilStatus.objects.get(id=1)
                        )

                    return HttpResponse("""
                                        <script>
                                            alert('La solicitud se ha actualizado correctamente.');
                                            window.location.href = '/coil-request/';
                                        </script>
                                    """)

            elif 'seleccionar' in request.POST and form.is_valid():
                marbetes_necesarios = int(request.POST.get('marbetes_necesarios', 0))
                try:
                    orden = coil_request_instance.FK_order_id
                    selected_order_coils = [int(id) for id in orden.coils.split(',')] if orden.coils else []

                    assigned_coil_ids = coil_request.objects.exclude(pk=pk).values_list('requested_coils', flat=True)
                    assigned_coil_ids = [int(id) for sublist in assigned_coil_ids for id in sublist.split(',') if id]

                    bobinas = coil.objects.filter(
                        id__in=selected_order_coils
                    )

                    total_marbetes = 0
                    selected_bobinas_ids = []
                    for bobina in bobinas:
                        if total_marbetes < marbetes_necesarios:
                            total_marbetes += bobina.missing
                            selected_bobinas_ids.append(bobina.id)

                    # Marcar bobinas según las seleccionadas y las preseleccionadas
                    selected_lote_coils = []
                    selected_quantity = 0
                    for bobina in bobinas:
                        if bobina.id in selected_bobinas_ids or bobina.id in previously_selected_coils:
                            selected_lote_coils.append(bobina.id)
                            selected_quantity += bobina.missing
                            bobina.selected = True
                        else:
                            bobina.selected = False

                        if selected_quantity >= marbetes_necesarios:
                            break

                    marbetes_totales = selected_quantity

                    return render(request, "cuervo/solicitudmarbete.html", {
                        "form": form,
                        "msg": msg,
                        "bobinas": bobinas,
                        "orden": orden,
                        "selected_lote_coils": selected_lote_coils,
                        "coil_request_instance": coil_request_instance,
                        "marbetes_necesarios": marbetes_necesarios,
                        "marbetes_totales": marbetes_totales
                    })

                except order.DoesNotExist:
                    msg = 'Orden de producción no encontrada'
                except lot.DoesNotExist:
                    msg = 'Lote granel no encontrado'
        else:
            if coil_request_instance.FK_order_id:
                orden = coil_request_instance.FK_order_id

            form = CreateCoilFormv2(initial={'ordenproduccion': coil_request_instance.FK_order_id.uniqueid})

            if orden:
                all_coils_in_order = coil.objects.filter(
                    id__in=[int(id) for id in orden.coils.split(',')] if orden.coils else []
                )

                assigned_coil_ids = coil_request.objects.exclude(pk=pk).values_list('requested_coils', flat=True)
                assigned_coil_ids = [int(id) for sublist in assigned_coil_ids for id in sublist.split(',') if id]

                bobinas = all_coils_in_order.exclude(id__in=assigned_coil_ids)

                selected_coil_ids = [int(id) for id in coil_request_instance.requested_coils.split(
                    ',')] if coil_request_instance.requested_coils else []
                selected_lote_coils = [bobina.id for bobina in all_coils_in_order.filter(id__in=selected_coil_ids)]

            return render(request, "cuervo/solicitudmarbete.html", {
                "form": form,
                "msg": msg,
                "bobinas": bobinas,
                "orden": orden,
                "selected_lote_coils": selected_lote_coils,
                "coil_request_instance": coil_request_instance,
                "marbetes_totales": marbetes_totales
            })
    except Exception as e:
        msg = f'Ocurrió un error: {str(e)}'
        return render(request, "cuervo/ErrorMsg.html", {
            "msg": msg
        })

codigos_escaneados = []
@permission_required('cuervo.view_labelstatus', login_url='/login/')
def init_label_damaged(request):
    return render(request, "cuervo/label_damaged.html", {'codigos': codigos_escaneados})

@permission_required('cuervo.view_labelstatus', login_url='/login/')
def agregar_codigo(request):
    if request.method == 'POST':
        codigo = request.POST['codigo']
        codigos_escaneados.append(codigo)  # Agregar el código a la lista interna
    return redirect('label-damaged')


def obtener_uniqueid(request):
    codigo = request.GET.get('url', '')
    print('ESTA ES LA URL/CODIGO:', codigo)
    try:
        if codigo.startswith('http'):
            # Buscar por URL
            uniqueid = label.objects.get(url=codigo).uniqueid
        else:
            # Buscar por uniqueid
            uniqueid = label.objects.get(uniqueid=codigo).uniqueid
        return JsonResponse({'uniqueid': uniqueid})
    except label.DoesNotExist:
        return JsonResponse({'error': 'Código no encontrado'}, status=404)


@permission_required('cuervo.view_labelstatus', login_url='/login/')
def confirmar_listado(request):
    codigos_input = request.POST.get('codigos', '')  # Obtener los códigos del campo oculto
    codigos_escaneados = codigos_input.split(',')  # Separar las entradas escaneadas

    labelStatus_daniado = labelStatus.objects.get(name="Dañados")

    # Crear listas para almacenar los cambios
    codigos_cambiados = []
    codigos_no_cambiados = []

    for codigo in codigos_escaneados:
        # Verifica si el código es una URL o un folio
        if codigo.startswith('http://') or codigo.startswith('https://'):
            # Buscar por URL
            label_obj = label.objects.filter(url=codigo).first()
        else:
            # Buscar por folio
            label_obj = label.objects.filter(uniqueid=codigo).first()

        if label_obj:
            if label_obj.FK_labelStatus_id != labelStatus_daniado:
                label_obj.FK_labelStatus_id = labelStatus_daniado  # Asigna la instancia directamente
                label_obj.save()
                codigos_cambiados.append(label_obj.uniqueid)
            else:
                codigos_no_cambiados.append(label_obj.uniqueid + ' Ya Dañado')
        else:
            codigos_no_cambiados.append(codigo + ' No existe')

    # Mensajes de éxito y advertencia
    if codigos_cambiados:
        messages.success(request, f"Los siguientes Marbetes se cambiaron de estado: {', '.join(codigos_cambiados)}")

    if codigos_no_cambiados:
        messages.warning(request, f"Los siguientes Marbetes no se cambiaron de estado: {', '.join(codigos_no_cambiados)}")

    return redirect('label-damaged')

@permission_required('cuervo.view_labelstatus', login_url='/login/')
def quitar_codigo(request, codigo):
    if codigo in codigos_escaneados:
        codigos_escaneados.remove(codigo)  # Remover el código de la lista interna
    return redirect('label-damaged')

@permission_required('cuervo.delete_coil_request', login_url='/login/')
def delete_coil_request(request, pk):
    coil_request_instance = get_object_or_404(coil_request, pk=pk)

    # Obtener las bobinas asociadas a la solicitud
    requested_coils_ids = [int(id) for id in coil_request_instance.requested_coils.split(',')] if coil_request_instance.requested_coils else []

    # Cambiar el estado de las bobinas a estatus 1
    coil.objects.filter(id__in=requested_coils_ids).update(FK_coilStatus_id=coilStatus.objects.get(id=1))

    # Eliminar la solicitud
    coil_request_instance.delete()

    return redirect(reverse('coil-request'))
